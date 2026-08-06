import os
import re
import json
import io
import requests
import pandas as pd
import streamlit as st
from pypdf import PdfReader
import openpyxl

st.set_page_config(
    page_title="Extractor Biopsias Mama - Unidad Radiología",
    page_icon="🩺",
    layout="wide"
)

st.title("🩺 Gestor de Biopsias de Mama - Extracción y Registro")
st.markdown("""
Esta aplicación procesa los informes PDF (`...RX`, `...INCIS`, `...ESCIS`), extrae la información estructurada
según las reglas médicas de la unidad y anexa o **actualiza** los datos en el Excel sin duplicar pacientes.
""")

default_key = ""
try:
    if "GEMINI_API_KEY" in st.secrets:
        default_key = st.secrets["GEMINI_API_KEY"]
except Exception:
    pass

st.sidebar.header("⚙️ Configuración")
api_key = st.sidebar.text_input("Ingresa tu clave de API Gemini:", value=default_key, type="password")
selected_year = st.sidebar.selectbox("Año de destino para la hoja:", ["2026", "2025", "2024"])

SYSTEM_INSTRUCTIONS = """
Eres un asistente médico especializado en radiología y senología mamaria. Tu función es extraer exactamente los campos indicados a partir de los informes PDF proporcionados.

SI NO HAY INFORME QUIRÚRGICO (ESCIS), DEJA LOS CAMPOS DE GANGLIO CENTINELA Y VACIAMIENTO AXILAR TOTALMENTE VACÍOS ("").

DEVUELVE LOS DATOS EN UN JSON CON LAS SIGUIENTES LLAVES EXACTAS:

- NH: Número de historia clínica del informe.
- NUHSA: Número de identificación andaluza de la paciente.
- NOMBRE: Nombre y apellidos en MAYÚSCULAS, sin comas.
- EDAD: Edad reflejada en el informe.
- FECHA_BIOPSIA: Fecha de realización/firma de la prueba radiológica (DD/MM/AAAA).
- SCREENING: Marcar "Sí" o "No" (PDPCM / cribado = "Sí").
- MAMOGRAFIAS_PREVIAS_1: Dejar VACÍO ("").
- MAMOGRAFIAS_PREVIAS_2: Dejar VACÍO ("").
- CLINICA: "Asintomática" o síntoma principal.
- ANTECEDENTES_MISMA_MAMA: "Sí" o "No".
- ANTECEDENTES_CONTRALATERAL: "Sí" o "No".
- ASPECTO_MAMOGRAFICO: Descriptor mamográfico (ej. "Lesión espiculada", "No mamografía").
- MAMOGRAFIA_CON_CONTRASTE: "Sí" o "No".
- BAV: "Sí" o "No".
- TAMAÑO_RX: Tamaño en CENTÍMETROS (cm). Ej: 18 mm -> 1.8 cm.
- MULTICENTRICO: "Sí" o "No".
- MULTIFOCAL: "Sí" o "No".
- BIRADS: Formato 'BIRADS X' (ej. 'BIRADS 5').
- RESULTADO_AP_MAMA: Resultado histológico con Grado (ej. 'Carcinoma ductal infiltrante G2').
- SUBCLASIFICACION_MOLECULAR: Marcadores moleculares (ej. 'Luminal A', 'Luminal B', 'HER2+', 'Triple Negativo').
- RM: "Sí", "No" o "No realizada".
- ECO_AXILA_ACTO_UNICO: "Sí" o "No".
- ESTADIO_ECOGRAFICO_AXILAR: N0, N1, N2.
- SOSPECHA_ECO_AXILAR: "Sí" o "No".
- PAAF_AXILA: "Sí" o "No".
- BAG_AXILA: "Sí" o "No".
- RESULTADO_BIOPSIA_AXILAR: "Sí" o "No".
- GANGLIO_CENTINELA: "Sí" o "No" (O "" si no hay ESCIS).
- RESULTADO_GANGLIO_CENTINELA: "Sí", "No", "MICROMETÁSTASIS" (O "" si no hay ESCIS).
- VACIAMIENTO_AXILAR: "Sí" o "No" (O "" si no hay ESCIS).
- RESULTADO_VACIAMIENTO: "Sí", "No", o VACÍO ("") si VACIAMIENTO_AXILAR es "No" o si no hay ESCIS.
- COMENTARIO: Dejar VACÍO ("") o nota relevante.

REGLAS ESTRUCTURALES DE SALIDA:
- Devuelve ÚNICAMENTE un objeto JSON válido.
"""

col1, col2 = st.columns(2)

with col1:
    st.markdown("### 1. Archivo Excel Registro (Opcional)")
    excel_file = st.file_uploader("Sube tu Excel máster existente (.xlsx):", type=["xlsx"])

with col2:
    st.markdown("### 2. Informes PDF de la paciente")
    uploaded_files = st.file_uploader("Sube los PDFs (RX, INCIS, ESCIS):", type=["pdf"], accept_multiple_files=True)

def get_available_models(clean_key):
    list_url = f"https://generativelanguage.googleapis.com/v1beta/models?key={clean_key}"
    try:
        res = requests.get(list_url)
        if res.status_code == 200:
            models_data = res.json().get("models", [])
            valid_models = []
            for m in models_data:
                methods = m.get("supportedGenerationMethods", [])
                if "generateContent" in methods:
                    name = m.get("name", "").replace("models/", "")
                    if "gemini" in name:
                        valid_models.append(name)
            return valid_models
    except Exception:
        pass
    return ["gemini-1.5-flash-latest", "gemini-2.0-flash", "gemini-1.5-pro-latest", "gemini-1.5-flash"]

def get_first_empty_row(ws, check_col=2, start_row=2):
    """Encuentra la primera fila vacía continua basándose en la columna B (NUHSA)."""
    r = start_row
    while True:
        val = ws.cell(r, check_col).value
        if val is None or str(val).strip() == "":
            return r
        r += 1

def find_existing_patient_row(ws, nuhsa_val, nh_val, max_search_row=150):
    """Busca si el paciente ya existe en la hoja comparando NUHSA o NH."""
    clean_nuhsa = str(nuhsa_val).strip() if nuhsa_val else ""
    clean_nh = str(nh_val).strip() if nh_val else ""

    for r in range(2, max_search_row):
        excel_nh = str(ws.cell(r, 1).value or "").strip()
        excel_nuhsa = str(ws.cell(r, 2).value or "").strip()

        if (clean_nuhsa and clean_nuhsa == excel_nuhsa) or (clean_nh and clean_nh == excel_nh):
            return r
    return None

if uploaded_files and api_key:
    if st.button("🚀 Procesar Documentos y Extraer Datos"):
        with st.spinner("Analizando informes médicos..."):
            try:
                clean_api_key = api_key.strip()
                texts = {}
                for f in uploaded_files:
                    reader = PdfReader(f)
                    text = "\n".join([page.extract_text() or "" for page in reader.pages])
                    texts[f.name] = text

                prompt_text = f"Extrae los datos de los siguientes informes médicos siguiendo estrictamente las instrucciones:\n\n{json.dumps(texts, ensure_ascii=False)}"

                model_candidates = get_available_models(clean_api_key)
                extracted_text = None

                for model_name in model_candidates:
                    for api_ver in ["v1beta", "v1"]:
                        url = f"https://generativelanguage.googleapis.com/{api_ver}/models/{model_name}:generateContent?key={clean_api_key}"
                        headers = {"Content-Type": "application/json"}
                        payload = {
                            "contents": [{"parts": [{"text": prompt_text}]}],
                            "systemInstruction": {"parts": [{"text": SYSTEM_INSTRUCTIONS}]},
                            "generationConfig": {"temperature": 0.0, "responseMimeType": "application/json"}
                        }

                        response = requests.post(url, headers=headers, json=payload)
                        res_json = response.json()

                        if response.status_code == 200 and "candidates" in res_json:
                            try:
                                extracted_text = res_json["candidates"][0]["content"]["parts"][0]["text"]
                                break
                            except (KeyError, IndexError):
                                continue
                    if extracted_text:
                        break

                if not extracted_text:
                    st.error("⚠️ No se pudo conectar con los modelos de IA.")
                    st.stop()

                st.session_state["result_json"] = extracted_text
                st.success("✅ ¡Datos extraídos correctamente!")

            except Exception as e:
                st.error(f"⚠️ Error al procesar: {str(e)}")

if "result_json" in st.session_state:
    st.markdown("---")
    st.subheader("📋 Previsualización y Edición del Caso Extraído")
    
    try:
        raw_json = st.session_state["result_json"]
        cleaned_json = re.sub(r"^```json\s*", "", raw_json, flags=re.MULTILINE)
        cleaned_json = re.sub(r"^```\s*", "", cleaned_json, flags=re.MULTILINE).strip()
        
        data = json.loads(cleaned_json)
        
        if isinstance(data, dict):
            new_df = pd.DataFrame([data])
        else:
            new_df = pd.DataFrame(data)
        
        target_column_order = [
            "NH", "NUHSA", "NOMBRE", "EDAD", "FECHA_BIOPSIA", "SCREENING",
            "MAMOGRAFIAS_PREVIAS_1", "MAMOGRAFIAS_PREVIAS_2", "CLINICA",
            "ANTECEDENTES_MISMA_MAMA", "ANTECEDENTES_CONTRALATERAL", "ASPECTO_MAMOGRAFICO",
            "MAMOGRAFIA_CON_CONTRASTE", "BAV", "TAMAÑO_RX", "MULTICENTRICO",
            "MULTIFOCAL", "BIRADS", "RESULTADO_AP_MAMA", "SUBCLASIFICACION_MOLECULAR",
            "RM", "ECO_AXILA_ACTO_UNICO", "ESTADIO_ECOGRAFICO_AXILAR", "SOSPECHA_ECO_AXILAR",
            "PAAF_AXILA", "BAG_AXILA", "RESULTADO_BIOPSIA_AXILAR", "GANGLIO_CENTINELA",
            "RESULTADO_GANGLIO_CENTINELA", "VACIAMIENTO_AXILAR", "RESULTADO_VACIAMIENTO",
            "COMENTARIO"
        ]

        for col in target_column_order:
            if col not in new_df.columns:
                new_df[col] = ""

        new_df["MAMOGRAFIAS_PREVIAS_1"] = ""
        new_df["MAMOGRAFIAS_PREVIAS_2"] = ""

        new_df = new_df[target_column_order]

        edited_df = st.data_editor(new_df, num_rows="dynamic", use_container_width=True)
        
        st.markdown("---")
        
        if excel_file is not None:
            excel_bytes = excel_file.getvalue()
            wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
            
            target_sheet_name = str(selected_year).strip()
            ws = wb[target_sheet_name] if target_sheet_name in wb.sheetnames else wb.create_sheet(title=target_sheet_name)
            
            updated_count = 0
            created_count = 0

            for _, row in edited_df.iterrows():
                nh_val = row.get("NH", "")
                nuhsa_val = row.get("NUHSA", "")

                # Comprobar si la paciente ya está registrada
                existing_row = find_existing_patient_row(ws, nuhsa_val, nh_val)

                if existing_row:
                    target_row = existing_row
                    updated_count += 1
                else:
                    target_row = get_first_empty_row(ws, check_col=2, start_row=2)
                    created_count += 1

                row_values = row.tolist()
                for col_idx, val in enumerate(row_values, start=1):
                    # Forzar vacías G (7) y H (8)
                    if col_idx in [7, 8]:
                        ws.cell(row=target_row, column=col_idx, value=None)
                    else:
                        clean_val = str(val).strip() if val is not None else ""
                        # Solo sobreescribir si trae un valor con contenido (o si es registro nuevo)
                        if clean_val != "" or not existing_row:
                            ws.cell(row=target_row, column=col_idx, value=clean_val if clean_val != "" else None)

            output_buffer = io.BytesIO()
            wb.save(output_buffer)
            final_excel = output_buffer.getvalue()

            if updated_count > 0:
                st.info(f"ℹ️ Se actualizó y completó la información de {updated_count} paciente(s) previamente registrada(s).")
            if created_count > 0:
                st.success(f"✅ Se creó el registro de {created_count} paciente(s) nueva(s).")

            st.download_button(
                label=f"📥 Confirmar y Descargar Excel Actualizado ({selected_year})",
                data=final_excel,
                file_name=f"Registro_Biopsias_Actualizado.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            csv = edited_df.to_csv(index=False).encode('utf-8')
            st.download_button(
                label="📥 Confirmar Visto Bueno y Descargar Caso como CSV",
                data=csv,
                file_name=f"caso_extraido_{selected_year}.csv",
                mime="text/csv"
            )

    except Exception as e:
        st.error(f"Error al procesar los datos: {e}")

    except Exception as e:
        st.error(f"Error al procesar los datos: {e}")
